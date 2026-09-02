from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

_TRANSACTION_ROOT = Path(__file__).resolve().parent.parent / "franchise" / "All Team Transactions BAA NBA history"

_SEASON_WORKBOOK_TEAMS: dict[int, dict[str, str]] = {
    1947: {
        "Boston_Celtics.xlsx": "Boston Celtics",
        "Chicago_Stags.xlsx": "Chicago Stags",
        "Cleveland_Rebels.xlsx": "Cleveland Rebels",
        "Detroit_Falcons.xlsx": "Detroit Falcons",
        "New_York_Knicks.xlsx": "New York Knicks",
        "Golden_State_Warriors.xlsx": "Philadelphia Warriors",
        "Pittsburgh_Ironmen.xlsx": "Pittsburgh Ironmen",
        "Providence_Steamrollers.xlsx": "Providence Steamrollers",
        "St._Louis_Bombers.xlsx": "St. Louis Bombers",
        "Toronto_Huskies.xlsx": "Toronto Huskies",
        "Washington_Capitols.xlsx": "Washington Capitols",
    }
}
_ADD_SHEETS = frozenset({"Signing", "Waiver Claim"})
_EVENT_SHEETS = frozenset((*_ADD_SHEETS, "Trade"))
_EVENT_PRIORITY = {"Signing": 0, "Waiver Claim": 1, "Trade": 2}
_BAA_NBA_LEAGUES = frozenset({"BAA", "NBA"})


@dataclass(frozen=True)
class TransactionEvent:
    event_date: date
    event_type: str
    team_name: str
    player_name: str
    source_file: str



@dataclass(frozen=True)
class TransactionTeamResolution:
    covered: bool
    matched: bool
    ambiguous: bool
    team_name: str | None
    event: TransactionEvent | None


@dataclass(frozen=True)
class _SeasonTransactionIndex:
    events_by_player_key: dict[str, tuple[TransactionEvent, ...]]
    ambiguous_player_keys: frozenset[str]



def resolve_baa_nba_transaction_team(
    *,
    season: int,
    player_name: str,
    source_league: str | None,
    source_leagues: tuple[str, ...] = (),
) -> TransactionTeamResolution:
    leagues = {
        str(league or "").strip().upper()
        for league in (source_league, *source_leagues)
        if str(league or "").strip()
    }
    if not leagues.intersection(_BAA_NBA_LEAGUES):
        return TransactionTeamResolution(False, False, False, None, None)
    if int(season) not in _SEASON_WORKBOOK_TEAMS:
        return TransactionTeamResolution(False, False, False, None, None)

    player_key = _identity(player_name)
    index = _season_transaction_index(int(season))
    if player_key in index.ambiguous_player_keys:
        return TransactionTeamResolution(True, False, True, None, None)

    events = index.events_by_player_key.get(player_key, ())
    if not events:
        return TransactionTeamResolution(True, False, False, None, None)
    event = events[-1]
    return TransactionTeamResolution(
        covered=True,
        matched=True,
        ambiguous=False,
        team_name=event.team_name,
        event=event,
    )


@lru_cache(maxsize=None)
def _season_transaction_index(season: int) -> _SeasonTransactionIndex:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("BAA/NBA transaction sorting requires openpyxl") from exc

    workbook_teams = _SEASON_WORKBOOK_TEAMS.get(int(season))
    if workbook_teams is None:
        return _SeasonTransactionIndex({}, frozenset())
    if not _TRANSACTION_ROOT.is_dir():
        raise FileNotFoundError(f"BAA/NBA transaction folder is missing: {_TRANSACTION_ROOT}")

    events_by_key: dict[str, list[TransactionEvent]] = {}
    raw_names_by_key: dict[str, set[str]] = {}
    seen_events: set[tuple[str, date, str, str]] = set()
    for filename, workbook_team in workbook_teams.items():
        path = _TRANSACTION_ROOT / filename
        if not path.is_file():
            raise FileNotFoundError(f"BAA/NBA transaction workbook is missing: {path}")
        workbook = load_workbook(path, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            if worksheet.title not in _EVENT_SHEETS:
                continue
            rows = worksheet.iter_rows(values_only=True)
            try:
                headers = tuple(str(value or "").strip() for value in next(rows))
            except StopIteration:
                continue
            for values in rows:
                row = dict(zip(headers, values))
                if _as_int(row.get("season")) != int(season):
                    continue
                player_name = str(row.get("player") or "").strip()
                if not player_name or row.get("date") in (None, ""):
                    continue
                if worksheet.title == "Trade":
                    team_name = str(row.get("team_1") or "").strip()
                    if not team_name:
                        continue
                else:
                    team_name = workbook_team
                player_key = _identity(player_name)
                event = TransactionEvent(
                    event_date=_as_date(row["date"]),
                    event_type=worksheet.title,
                    team_name=team_name,
                    player_name=player_name,
                    source_file=filename,
                )
                event_identity = (player_key, event.event_date, event.event_type, event.team_name)
                if event_identity in seen_events:
                    continue
                seen_events.add(event_identity)
                events_by_key.setdefault(player_key, []).append(event)
                raw_names_by_key.setdefault(player_key, set()).add(player_name)

    ordered = {
        player_key: tuple(
            sorted(
                events,
                key=lambda event: (
                    event.event_date,
                    _EVENT_PRIORITY[event.event_type],
                    event.team_name,
                    event.source_file,
                ),
            )
        )
        for player_key, events in events_by_key.items()
    }
    ambiguous = frozenset(player_key for player_key, names in raw_names_by_key.items() if len(names) != 1)
    return _SeasonTransactionIndex(ordered, ambiguous)


def _identity(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    ascii_text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Z0-9]+", "", ascii_text.upper())


def _as_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.fromisoformat(str(value)).date()


def _as_int(value: Any) -> int:
    return int(value or 0)
