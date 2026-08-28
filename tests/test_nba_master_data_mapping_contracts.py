from __future__ import annotations

import ast
import hashlib
import json
import sys
from collections import Counter
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MAPPINGS_DIR = ROOT / "nba2k_editor" / "Player Generator" / "mappings"
if str(MAPPINGS_DIR) not in sys.path:
    sys.path.insert(0, str(MAPPINGS_DIR))

from mapping_contracts import (  # type: ignore[import-not-found]  # noqa: E402
    CONTRACT_DATA_PATH,
    CONTRACTS,
    REGISTRY,
    EvidenceRole,
    RelationshipStatus,
    SqlColumnRole,
    SqlColumnStatus,
    active_2k26_destinations,
    validate_registry,
)

WORKBOOK = MAPPINGS_DIR / "NBAMASTERDATAMAPPINGS.xlsx"
EXPECTED_WORKBOOK_SHA256 = "40a308ee21ef5d3849d91cd32e5e7fb2ec13f17083d5047fab84662286a33025"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _split_lines(value: object) -> tuple[str, ...]:
    return tuple(line for line in str(value or "").splitlines() if line)


def test_registry_keys_are_qualified_and_contract_module_has_no_substring_dispatch() -> None:
    qualified_keys = set(REGISTRY.by_qualified_name())
    assert qualified_keys == {contract.qualified_name for contract in CONTRACTS}
    assert all(" | " in key for key in qualified_keys)
    assert not any(contract.header in qualified_keys for contract in CONTRACTS)

    tree = ast.parse((MAPPINGS_DIR / "mapping_contracts.py").read_text(encoding="utf-8"))
    forbidden_calls = {
        node.func.attr
        for node in ast.walk(tree)
        if isinstance(node, ast.Call)
        and isinstance(node.func, ast.Attribute)
        and node.func.attr in {"startswith", "endswith", "find"}
    }
    assert forbidden_calls == set()


def test_sql_only_shotquality_source_is_explicitly_gated_and_fully_classified() -> None:
    assert len(REGISTRY.sql_only_sources) == 1
    source = REGISTRY.sql_only_sources[0]
    assert source.table == "crafted_source_shotquality"
    assert source.workbook_page == "ShotQuality"
    assert source.included_in_workbook is False
    assert source.provenance_database == "NBA_DATA_Master.sqlite"
    assert source.provenance_audit_key == "file:crafted_source_shotquality"
    assert source.identity_source_columns == ("year", "nba_id")
    assert source.identity_map_table == "crafted_player_id_map"
    assert source.identity_map_source_column == "nba_id"
    assert source.identity_map_target_column == "player_id"
    assert source.identity_map_status_column == "status"
    assert source.identity_map_status_value == "mapped"
    assert source.allow_name_fallback is False
    assert source.season_coverage == (2026,)
    assert source.expected_row_count == 673
    assert source.expected_duplicate_identity_count == 0
    assert tuple(column.name for column in source.columns) == (
        "nba_id",
        "player_name",
        "team_abbreviation",
        "year",
        "ots",
        "osq",
        "omake",
        "oft",
        "dts",
        "dsq",
        "dcontest",
        "dft",
        "cts",
        "csq",
        "cmake",
        "cft",
        "off_poss",
        "source_row_id",
    )
    assert all(column.status is SqlColumnStatus.REVIEW for column in source.columns)
    assert all(not column.approved_destinations for column in source.columns)
    by_name = {column.name: column for column in source.columns}
    assert by_name["nba_id"].role is SqlColumnRole.IDENTITY
    assert by_name["year"].role is SqlColumnRole.IDENTITY
    assert by_name["source_row_id"].role is SqlColumnRole.AUDIT
    assert by_name["player_name"].role is SqlColumnRole.DISPLAY
    assert by_name["dcontest"].role is SqlColumnRole.METRIC
    assert by_name["dcontest"].direction == "unproven"


def test_sql_only_gate_rejects_name_identity_and_unclassified_inclusion() -> None:
    source = REGISTRY.sql_only_sources[0]
    with pytest.raises(ValueError, match="permits name fallback"):
        validate_registry(
            replace(
                REGISTRY,
                sql_only_sources=(replace(source, allow_name_fallback=True),),
            )
        )
    with pytest.raises(ValueError, match="not fully classified"):
        validate_registry(
            replace(
                REGISTRY,
                sql_only_sources=(replace(source, included_in_workbook=True),),
            )
        )


def test_registry_reproduces_verified_baseline_counts() -> None:
    assert _sha256(WORKBOOK) == EXPECTED_WORKBOOK_SHA256
    assert REGISTRY.baseline_sha256 == EXPECTED_WORKBOOK_SHA256
    assert REGISTRY.source_count == 549
    assert REGISTRY.edge_count == 1246
    assert len(CONTRACTS) == 549
    assert len(REGISTRY.by_qualified_name()) == 549


def test_contract_data_is_static_exact_source_inventory() -> None:
    payload = json.loads(CONTRACT_DATA_PATH.read_text(encoding="utf-8"))
    assert payload["schema_version"] == 1
    assert payload["baseline"]["source_count"] == 549
    assert payload["baseline"]["edge_count"] == 1246
    assert len(payload["sources"]) == 549
    assert all(source["page"] and source["header"] for source in payload["sources"])
    assert all("qualified" not in source for source in payload["sources"])


def test_registry_cannot_collapse_to_bare_headers_or_substring_dispatch() -> None:
    header_counts = Counter(contract.header for contract in CONTRACTS)
    assert any(count > 1 for count in header_counts.values())
    assert len(REGISTRY.by_qualified_name()) == len(CONTRACTS)
    assert all(" | " in key for key in REGISTRY.by_qualified_name())

    module_source = (MAPPINGS_DIR / "mapping_contracts.py").read_text(encoding="utf-8")
    for forbidden_dispatch in ("startswith(", "endswith(", "re.compile", "fnmatch"):
        assert forbidden_dispatch not in module_source


def test_every_relationship_uses_an_active_exact_destination() -> None:
    active = set(active_2k26_destinations())
    assert len(active) == 206
    mapped = {
        relationship.destination
        for contract in CONTRACTS
        for relationship in contract.relationships
    }
    assert mapped <= active
    assert len(mapped) == 191


def test_registry_matches_persisted_source_ledger_exactly() -> None:
    workbook = load_workbook(WORKBOOK, data_only=False, read_only=False)
    ledger = workbook["Source Ledger"]
    actual: dict[str, dict[str, object]] = {}
    for row in range(2, ledger.max_row + 1):
        header = str(ledger.cell(row, 1).value or "")
        page = str(ledger.cell(row, 2).value or "")
        qualified = f"{page} | {header}"
        actual[qualified] = {
            "source_level": str(ledger.cell(row, 3).value or ""),
            "treatment": str(ledger.cell(row, 4).value or ""),
            "destinations": _split_lines(ledger.cell(row, 5).value),
            "review_reason": ledger.cell(row, 6).value,
        }
    workbook.close()

    assert set(actual) == set(REGISTRY.by_qualified_name())
    for qualified, contract in REGISTRY.by_qualified_name().items():
        row = actual[qualified]
        assert row["source_level"] == contract.source_level
        assert row["treatment"] == contract.treatment
        assert row["destinations"] == contract.destinations
        assert row["review_reason"] == contract.review_reason


def test_registry_matches_persisted_forward_mapping_exactly() -> None:
    workbook = load_workbook(WORKBOOK, data_only=False, read_only=False)
    headers = workbook["Field Headers"]
    actual: dict[str, tuple[str, ...]] = {}
    for col in range(2, headers.max_column + 1):
        destination = headers.cell(1, col).value
        if destination:
            actual[str(destination)] = tuple(
                str(headers.cell(row, col).value)
                for row in range(4, headers.max_row + 1)
                if headers.cell(row, col).value
            )
    workbook.close()

    expected = REGISTRY.sources_by_destination()
    assert actual == {
        destination: expected.get(destination, ())
        for destination in actual
    }


def test_literal_reviews_are_unmapped_and_explained() -> None:
    literal_reviews = [contract for contract in CONTRACTS if contract.treatment == "Review"]
    nonreviews = [contract for contract in CONTRACTS if contract.treatment != "Review"]
    assert len(literal_reviews) == 194
    assert len(nonreviews) == 355
    assert all(not contract.relationships for contract in literal_reviews)
    assert all(contract.review_reason for contract in literal_reviews)
    assert all(contract.evidence_role is EvidenceRole.REVIEW for contract in literal_reviews)
    assert all(contract.relationship_status is RelationshipStatus.REVIEW for contract in literal_reviews)
    assert all(contract.relationships for contract in nonreviews)


def test_generalized_sources_are_not_claimed_as_exact() -> None:
    generalized = [
        contract
        for contract in CONTRACTS
        if contract.treatment == "Generalized formula context/proxy"
    ]
    assert len(generalized) == 33
    assert all(contract.relationship_status is RelationshipStatus.GENERALIZED_PROXY for contract in generalized)
    assert all(
        relationship.relationship_status is RelationshipStatus.GENERALIZED_PROXY
        for contract in generalized
        for relationship in contract.relationships
    )


def test_dunk_makes_are_the_approved_attribute_tendency_exception() -> None:
    contract = REGISTRY.by_qualified_name()["Player Shooting | num_of_dunks"]
    assert contract.relationship_status is RelationshipStatus.APPROVED_EXCEPTION
    assert set(contract.destinations) == {
        "Attributes/DRIVINGDUNK",
        "Tendencies/DRIVINGDUNK",
        "Tendencies/STANDINGDUNK",
    }
    assert all(
        relationship.relationship_status is RelationshipStatus.APPROVED_EXCEPTION
        for relationship in contract.relationships
    )


def test_historical_shot_and_touches_fallbacks_are_explicit() -> None:
    for contract in CONTRACTS:
        for relationship in contract.relationships:
            if relationship.evidence_role is not EvidenceRole.HISTORICAL_FALLBACK:
                continue
            assert relationship.relationship_status is RelationshipStatus.APPROVED_EXCEPTION
            assert relationship.destination in {"Tendencies/SHOT", "Tendencies/TOUCHES"}
            assert "fallback" in relationship.rationale.casefold()


def test_two_and_three_point_makes_do_not_author_shot_tendencies() -> None:
    exact_make_sources = {
        "Team Stats Per 100 Possessions | x3p_per_100_poss",
        "Team Stats Per 100 Possessions | x2p_per_100_poss",
        "Team Stats Per Game | x3p_per_game",
        "Team Stats Per Game | x2p_per_game",
        "Team Totals | x3p",
        "Team Totals | x2p",
        "Player Per Game | x3p_per_game",
        "Player Per Game | x2p_per_game",
        "Player Per 36 min | x3p_per_36_min",
        "Player Per 36 min | x2p_per_36_min",
        "Player Per 100 Poss | x3p_per_100_poss",
        "Player Per 100 Poss | x2p_per_100_poss",
        "Player Totals | x3p",
        "Player Totals | x2p",
    }
    contracts = REGISTRY.by_qualified_name()
    assert exact_make_sources <= set(contracts)
    for source in exact_make_sources:
        assert not {
            destination
            for destination in contracts[source].destinations
            if destination.startswith("Tendencies/")
        }, source


def test_exact_attempt_and_range_share_sources_do_not_author_shooting_attributes() -> None:
    exact_attempt_sources = {
        "Team Stats Per 100 Possessions | fga_per_100_poss",
        "Team Stats Per 100 Possessions | x3pa_per_100_poss",
        "Team Stats Per 100 Possessions | x2pa_per_100_poss",
        "Team Stats Per Game | fga_per_game",
        "Team Stats Per Game | x3pa_per_game",
        "Team Stats Per Game | x2pa_per_game",
        "Team Totals | fga",
        "Team Totals | x3pa",
        "Team Totals | x2pa",
        "Player Per Game | fga_per_game",
        "Player Per Game | x3pa_per_game",
        "Player Per Game | x2pa_per_game",
        "Player Per 36 min | fga_per_36_min",
        "Player Per 36 min | x3pa_per_36_min",
        "Player Per 36 min | x2pa_per_36_min",
        "Player Per 100 Poss | fga_per_100_poss",
        "Player Per 100 Poss | x3pa_per_100_poss",
        "Player Per 100 Poss | x2pa_per_100_poss",
        "Player Totals | fga",
        "Player Totals | x3pa",
        "Player Totals | x2pa",
        "Team Summaries | x3p_ar",
        "Player Shooting | percent_fga_from_x2p_range",
        "Player Shooting | percent_fga_from_x0_3_range",
        "Player Shooting | percent_fga_from_x3_10_range",
        "Player Shooting | percent_fga_from_x10_16_range",
        "Player Shooting | percent_fga_from_x16_3p_range",
        "Player Shooting | percent_fga_from_x3p_range",
        "Player Advanced | x3p_ar",
    }
    contracts = REGISTRY.by_qualified_name()
    assert exact_attempt_sources <= set(contracts)
    for source in exact_attempt_sources:
        assert not {
            destination
            for destination in contracts[source].destinations
            if destination.startswith("Attributes/")
        }, source


def test_free_throw_percentage_and_attempt_ownership_is_separate() -> None:
    contracts = REGISTRY.by_qualified_name()
    percentage_sources = {
        "Player Per Game | ft_percent",
        "Player Per 36 min | ft_percent",
        "Player Per 100 Poss | ft_percent",
        "Player Totals | ft_percent",
    }
    attempt_sources = {
        "Player Per Game | fta_per_game",
        "Player Per 36 min | fta_per_36_min",
        "Player Per 100 Poss | fta_per_100_poss",
        "Player Totals | fta",
    }
    assert set(REGISTRY.sources_by_destination()["Attributes/FREETHROW"]) == percentage_sources
    for source in percentage_sources:
        assert "Attributes/FREETHROW" in contracts[source].destinations
        relationship = next(
            relationship
            for relationship in contracts[source].relationships
            if relationship.destination == "Attributes/FREETHROW"
        )
        assert relationship.evidence_role is EvidenceRole.DIRECT
        assert relationship.relationship_status is RelationshipStatus.EXACT
        assert relationship.decision_id == "RULE_FT_PERCENT"
    for source in attempt_sources:
        assert "Attributes/FREETHROW" not in contracts[source].destinations
        assert "Attributes/DRAWFOUL" in contracts[source].destinations


def test_fgm_and_points_historical_fallback_edges_are_exactly_marked() -> None:
    contracts = REGISTRY.by_qualified_name()
    player_fallback_edges = {
        "Tendencies/TOUCHES": {
            "Player Per Game | fg_per_game",
            "Player Per 36 min | fg_per_36_min",
            "Player Per 100 Poss | fg_per_100_poss",
            "Player Totals | fg",
            "Player Per Game | fta_per_game",
            "Player Per 36 min | fta_per_36_min",
            "Player Per 100 Poss | fta_per_100_poss",
            "Player Totals | fta",
        },
        "Tendencies/SHOT": {
            "Player Per Game | pts_per_game",
            "Player Per 36 min | pts_per_36_min",
            "Player Per 100 Poss | pts_per_100_poss",
            "Player Totals | pts",
        },
    }
    team_denominator_edges = {
        "Tendencies/TOUCHES": {
            "Team Stats Per 100 Possessions | fg_per_100_poss",
            "Team Stats Per Game | fg_per_game",
            "Team Totals | fg",
            "Team Stats Per 100 Possessions | fta_per_100_poss",
            "Team Stats Per Game | fta_per_game",
            "Team Totals | fta",
        },
        "Tendencies/SHOT": {
            "Team Stats Per 100 Possessions | pts_per_100_poss",
            "Team Stats Per Game | pts_per_game",
            "Team Totals | pts",
        },
    }
    for destination, sources in player_fallback_edges.items():
        for source in sources:
            relationship = next(
                relationship
                for relationship in contracts[source].relationships
                if relationship.destination == destination
            )
            assert relationship.evidence_role is EvidenceRole.HISTORICAL_FALLBACK
            assert relationship.relationship_status is RelationshipStatus.APPROVED_EXCEPTION
            assert relationship.decision_id
            assert relationship.conditions == (
                ("FGA_UNAVAILABLE", "AST_UNAVAILABLE", "USG_UNAVAILABLE", "EXACT_TEAM_SHARE")
                if destination == "Tendencies/TOUCHES"
                else ("FGA_UNAVAILABLE", "EXACT_TEAM_SHARE")
            )
    for destination, sources in team_denominator_edges.items():
        for source in sources:
            relationship = next(
                relationship
                for relationship in contracts[source].relationships
                if relationship.destination == destination
            )
            assert relationship.evidence_role is EvidenceRole.DENOMINATOR
            assert relationship.relationship_status is RelationshipStatus.APPROVED_EXCEPTION
            assert relationship.decision_id
            assert relationship.conditions == (
                ("FGA_UNAVAILABLE", "AST_UNAVAILABLE", "USG_UNAVAILABLE", "EXACT_TEAM_SHARE")
                if destination == "Tendencies/TOUCHES"
                else ("FGA_UNAVAILABLE", "EXACT_TEAM_SHARE")
            )


def test_contest_shot_current_sources_are_only_opponent_attempt_context() -> None:
    expected_sources = {
        "Opponent Stats Per 100 Possessions | opp_fga_per_100_poss",
        "Opponent Stats Per 100 Possessions | opp_x3pa_per_100_poss",
        "Opponent Stats Per 100 Possessions | opp_x2pa_per_100_poss",
        "Opponent Stats Per Game | opp_fga_per_game",
        "Opponent Stats Per Game | opp_x3pa_per_game",
        "Opponent Stats Per Game | opp_x2pa_per_game",
        "Opponent Totals | opp_fga",
        "Opponent Totals | opp_x3pa",
        "Opponent Totals | opp_x2pa",
    }
    assert set(REGISTRY.sources_by_destination()["Tendencies/CONTESTSHOT"]) == expected_sources
    contracts = REGISTRY.by_qualified_name()
    for source in expected_sources:
        relationship = next(
            item
            for item in contracts[source].relationships
            if item.destination == "Tendencies/CONTESTSHOT"
        )
        assert relationship.evidence_role is EvidenceRole.CONTEXT
        assert relationship.relationship_status is RelationshipStatus.BASELINE_UNREVIEWED


def test_heaves_remain_review_only() -> None:
    contracts = REGISTRY.by_qualified_name()
    for source in {
        "Player Shooting | num_heaves_attempted",
        "Player Shooting | num_heaves_made",
    }:
        contract = contracts[source]
        assert contract.treatment == "Review"
        assert contract.destinations == ()
        assert contract.relationship_status is RelationshipStatus.REVIEW


def test_split_rebound_attributes_have_independently_owned_evidence() -> None:
    by_destination = REGISTRY.sources_by_destination()
    offensive = set(by_destination["Attributes/OFFENSIVEREBOUND"])
    defensive = set(by_destination["Attributes/DEFENSEREBOUND"])
    assert {
        "Player Per Game | orb_per_game",
        "Player Per 36 min | orb_per_36_min",
        "Player Per 100 Poss | orb_per_100_poss",
        "Player Totals | orb",
        "Player Advanced | orb_percent",
    } <= offensive
    assert {
        "Player Per Game | drb_per_game",
        "Player Per 36 min | drb_per_36_min",
        "Player Per 100 Poss | drb_per_100_poss",
        "Player Totals | drb",
        "Player Advanced | drb_percent",
    } <= defensive
    assert offensive - defensive
    assert defensive - offensive
    contracts = REGISTRY.by_qualified_name()
    dual_total_sources = {
        "Player Per Game | trb_per_game",
        "Player Per 36 min | trb_per_36_min",
        "Player Per 100 Poss | trb_per_100_poss",
        "Player Totals | trb",
        "Player Advanced | trb_percent",
    }
    for source in dual_total_sources:
        rebound_relationships = [
            relationship
            for relationship in contracts[source].relationships
            if relationship.destination in {
                "Attributes/OFFENSIVEREBOUND",
                "Attributes/DEFENSEREBOUND",
            }
        ]
        assert len(rebound_relationships) == 2
        assert all(
            relationship.relationship_status is RelationshipStatus.BASELINE_UNREVIEWED
            and relationship.evidence_role is EvidenceRole.UNREVIEWED
            for relationship in rebound_relationships
        )
