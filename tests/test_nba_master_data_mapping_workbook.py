from __future__ import annotations

import shutil
import sys
from dataclasses import replace
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

ROOT = Path(__file__).resolve().parents[1]
MAPPINGS_DIR = ROOT / "nba2k_editor" / "Player Generator" / "mappings"
if str(MAPPINGS_DIR) not in sys.path:
    sys.path.insert(0, str(MAPPINGS_DIR))

from build_nba_master_data_mappings import (  # type: ignore[import-not-found]  # noqa: E402
    DEFAULT_DATABASE,
    EXPECTED_SHEETS,
    PAGE_TABLES,
    render_workbook,
    validate_sql_only_source_inventory,
    validate_sqlite_source_inventory,
)
from audit_nba_master_data_mappings import audit_workbook  # type: ignore[import-not-found]  # noqa: E402
from mapping_contracts import REGISTRY  # type: ignore[import-not-found]  # noqa: E402

WORKBOOK = MAPPINGS_DIR / "NBAMASTERDATAMAPPINGS.xlsx"


def _cell_signature(workbook_path: Path) -> dict[str, list[tuple[object, ...]]]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        return {
            sheet.title: [
                tuple(
                    (
                        cell.value,
                        cell.style_id,
                        None if cell.comment is None else (cell.comment.text, cell.comment.author),
                        None if cell.hyperlink is None else cell.hyperlink.target,
                    )
                    for cell in row
                )
                for row in sheet.iter_rows()
            ]
            for sheet in workbook.worksheets
        }
    finally:
        workbook.close()


def _structure_signature(workbook_path: Path) -> dict[str, object]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        return {
            "sheetnames": workbook.sheetnames,
            "defined_names": sorted((name, str(item)) for name, item in workbook.defined_names.items()),
            "sheets": {
                sheet.title: {
                    "dimension": sheet.calculate_dimension(),
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                    "auto_filter": sheet.auto_filter.ref,
                    "merges": sorted(str(item) for item in sheet.merged_cells.ranges),
                    "tables": sorted(
                        (
                            name,
                            sheet.tables[name].ref,
                            (
                                sheet.tables[name].tableStyleInfo.name
                                if sheet.tables[name].tableStyleInfo
                                else None
                            ),
                        )
                        for name in sheet.tables
                    ),
                    "validations": sorted(
                        (str(item.sqref), item.type, item.formula1, item.formula2)
                        for item in sheet.data_validations.dataValidation
                    ),
                    "conditional_formatting": sorted(
                        str(item.sqref) for item in sheet.conditional_formatting
                    ),
                    "charts": len(sheet._charts),
                    "images": len(sheet._images),
                    "state": sheet.sheet_state,
                }
                for sheet in workbook.worksheets
            },
        }
    finally:
        workbook.close()


def test_page_table_ownership_is_explicit_and_complete() -> None:
    assert len(PAGE_TABLES) == 23
    assert set(PAGE_TABLES) == {contract.page for contract in REGISTRY.contracts}
    validate_sqlite_source_inventory(DEFAULT_DATABASE)


def test_sql_only_source_inventory_matches_read_only_data_master() -> None:
    validate_sql_only_source_inventory(DEFAULT_DATABASE)


def test_sql_only_source_inventory_rejects_unclassified_column() -> None:
    source = REGISTRY.sql_only_sources[0]
    incomplete_registry = replace(
        REGISTRY,
        sql_only_sources=(replace(source, columns=source.columns[:-1]),),
    )
    with pytest.raises(ValueError, match="SQL-only column definitions differ"):
        validate_sql_only_source_inventory(DEFAULT_DATABASE, incomplete_registry)


def test_sql_only_source_inventory_rejects_provenance_drift() -> None:
    source = REGISTRY.sql_only_sources[0]
    drifted_registry = replace(
        REGISTRY,
        sql_only_sources=(replace(source, provenance_source_path="untrusted.csv"),),
    )
    with pytest.raises(ValueError, match="SQL-only provenance differs"):
        validate_sql_only_source_inventory(DEFAULT_DATABASE, drifted_registry)


def test_renderer_refuses_in_place_write() -> None:
    with pytest.raises(ValueError, match="refusing in-place"):
        render_workbook(template_path=WORKBOOK, output_path=WORKBOOK)


def test_renderer_reproduces_current_workbook_logically(tmp_path: Path) -> None:
    output = tmp_path / "rebuilt.xlsx"
    render_workbook(template_path=WORKBOOK, output_path=output)

    assert output.is_file()
    assert _cell_signature(output) == _cell_signature(WORKBOOK)
    assert _structure_signature(output) == _structure_signature(WORKBOOK)
    with ZipFile(output) as archive:
        assert archive.testzip() is None
        output_members = set(archive.namelist())
    with ZipFile(WORKBOOK) as archive:
        assert archive.testzip() is None
        template_members = set(archive.namelist())
    assert output_members == template_members
    persisted = load_workbook(output, read_only=False)
    try:
        assert persisted.sheetnames == EXPECTED_SHEETS
    finally:
        persisted.close()
    strict_result = audit_workbook(workbook_path=output)
    assert strict_result["success"] is False
    assert "workbook_sha256" in {
        violation["kind"] for violation in strict_result["violations"]
    }
    assert audit_workbook(workbook_path=output, strict_baseline_sha=False)["success"] is True


def test_permanent_audit_accepts_verified_workbook() -> None:
    result = audit_workbook(workbook_path=WORKBOOK)
    assert result["success"] is True
    assert result["counts"] == {
        "sources": 549,
        "nonreview_sources": 355,
        "review_sources": 194,
        "edges": 1246,
        "active_destinations": 206,
        "mapped_destinations": 191,
        "blank_destinations": 15,
    }
    assert result["violations"] == []
    assert result["source_labels"] == [contract.qualified_name for contract in REGISTRY.contracts]
    assert len(result["destination_labels"]) == 206


def test_permanent_audit_rejects_reverse_mapping_tamper(tmp_path: Path) -> None:
    tampered = tmp_path / "tampered.xlsx"
    shutil.copy2(WORKBOOK, tampered)
    workbook = load_workbook(tampered, data_only=False, read_only=False)
    try:
        ledger = workbook["Source Ledger"]
        row = next(
            row
            for row in range(2, ledger.max_row + 1)
            if ledger.cell(row, 1).value == "num_of_dunks"
            and ledger.cell(row, 2).value == "Player Shooting"
        )
        mapping_cell = ledger.cell(row, 5)
        assert isinstance(mapping_cell, Cell)
        mapping_cell.value = "Attributes/DRIVINGDUNK"
        workbook.save(tampered)
    finally:
        workbook.close()

    result = audit_workbook(workbook_path=tampered)
    assert result["success"] is False
    kinds = {violation["kind"] for violation in result["violations"]}
    assert "source_ledger_contract" in kinds
    assert "forward_reverse_symmetry" in kinds


def test_renderer_does_not_publish_before_candidate_verification(tmp_path: Path) -> None:
    template = tmp_path / "tampered-template.xlsx"
    output = tmp_path / "must-not-exist.xlsx"
    shutil.copy2(WORKBOOK, template)
    workbook = load_workbook(template, data_only=False, read_only=False)
    try:
        workbook.properties.creator = "tampered creator"
        workbook.save(template)
    finally:
        workbook.close()

    with pytest.raises(ValueError, match="content differs from registry baseline"):
        render_workbook(template_path=template, output_path=output)
    assert not output.exists()


@pytest.mark.parametrize("sheet_name", EXPECTED_SHEETS)
def test_permanent_audit_returns_structured_failure_for_missing_sheet(
    tmp_path: Path,
    sheet_name: str,
) -> None:
    tampered = tmp_path / f"missing-{sheet_name}.xlsx"
    shutil.copy2(WORKBOOK, tampered)
    workbook = load_workbook(tampered, data_only=False, read_only=False)
    try:
        workbook.remove(workbook[sheet_name])
        workbook.save(tampered)
    finally:
        workbook.close()

    result = audit_workbook(workbook_path=tampered, strict_baseline_sha=False)
    assert result["success"] is False
    assert {violation["kind"] for violation in result["violations"]} == {"sheet_order"}


def test_permanent_audit_returns_structured_failure_for_corrupt_zip(tmp_path: Path) -> None:
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not an xlsx zip")
    result = audit_workbook(workbook_path=corrupt)
    assert result["success"] is False
    assert result["zip_integrity"] == "failed"
    assert {violation["kind"] for violation in result["violations"]} == {
        "xlsx_zip_integrity"
    }


def test_permanent_audit_rejects_workbook_property_tamper_in_non_strict_mode(
    tmp_path: Path,
) -> None:
    tampered = tmp_path / "creator-tampered.xlsx"
    shutil.copy2(WORKBOOK, tampered)
    workbook = load_workbook(tampered, data_only=False, read_only=False)
    try:
        workbook.properties.creator = "tampered creator"
        workbook.save(tampered)
    finally:
        workbook.close()
    result = audit_workbook(workbook_path=tampered, strict_baseline_sha=False)
    assert "workbook_content" in {
        violation["kind"] for violation in result["violations"]
    }


def test_permanent_audit_rejects_notes_tamper_in_non_strict_mode(tmp_path: Path) -> None:
    tampered = tmp_path / "notes-tampered.xlsx"
    shutil.copy2(WORKBOOK, tampered)
    workbook = load_workbook(tampered, data_only=False, read_only=False)
    try:
        workbook["Notes"]["A1"] = "tampered note"
        workbook.save(tampered)
    finally:
        workbook.close()
    result = audit_workbook(workbook_path=tampered, strict_baseline_sha=False)
    kinds = {violation["kind"] for violation in result["violations"]}
    assert "workbook_content" in kinds
    assert "notes_contract" in kinds


def test_permanent_audit_rejects_whitespace_only_mapping_tamper(
    tmp_path: Path,
) -> None:
    tampered = tmp_path / "whitespace-tampered.xlsx"
    shutil.copy2(WORKBOOK, tampered)
    workbook = load_workbook(tampered, data_only=False, read_only=False)
    try:
        ledger = workbook["Source Ledger"]
        mapping_cell = next(
            ledger.cell(row, 5)
            for row in range(2, ledger.max_row + 1)
            if ledger.cell(row, 5).value
        )
        assert isinstance(mapping_cell, Cell)
        mapping_cell.value = f"{mapping_cell.value}\n"
        workbook.save(tampered)
    finally:
        workbook.close()
    result = audit_workbook(workbook_path=tampered, strict_baseline_sha=False)
    assert "workbook_content" in {
        violation["kind"] for violation in result["violations"]
    }


def test_permanent_audit_rejects_opaque_package_part_tamper(tmp_path: Path) -> None:
    tampered = tmp_path / "styles-tampered.xlsx"
    rewritten = tmp_path / "styles-tampered-rewritten.xlsx"
    shutil.copy2(WORKBOOK, tampered)
    with ZipFile(tampered) as archive:
        members = {name: archive.read(name) for name in archive.namelist()}
    styles = members["xl/styles.xml"]
    assert b"</styleSheet>" in styles
    members["xl/styles.xml"] = styles.replace(
        b"</styleSheet>",
        b"<!-- audit tamper -->\n</styleSheet>",
        1,
    )
    with ZipFile(rewritten, "w", compression=ZIP_DEFLATED) as archive:
        for name, content in members.items():
            archive.writestr(name, content)
    rewritten.replace(tampered)

    result = audit_workbook(workbook_path=tampered, strict_baseline_sha=False)
    assert "package_opaque_parts" in {
        violation["kind"] for violation in result["violations"]
    }
